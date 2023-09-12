#-----------------------------------------------------------------------------------------------------------------------------------------------------------------------
#Autor del codigo: Benjamin Ross
#
#   El objetivo de este .py es facilitar a las personas la eleccion de la secciones en la toma de ramos, utilizando el metodo de optimizacion MILP para minimizar la
#cantidad de ventanas dentro de un horario, maximizar los dias libres y los dias donde se sale antes de almuerzo.
#   Este proyecto fue hecho por amor arte, es decir, no esta autorizadaa su venta sin autorizacion explicita
#-----------------------------------------------------------------------------------------------------------------------------------------------------------------------

try:
    import requests
except ImportError:
    print("El módulo requests no está instalado. Instalando...")
    import subprocess
    subprocess.check_call(["pip", "install", "requests"])
    import requests
try:
    from bs4 import BeautifulSoup
except ImportError:
    print("El módulo BeautifulSoup no está instalado. Instalando...")
    import subprocess
    subprocess.check_call(["pip", "install", "beautifulsoup4"])
    from bs4 import BeautifulSoup
try:
    from pulp import *
except ImportError:
    print("El módulo pulp no está instalado. Instalando...")
    import subprocess
    subprocess.check_call(["pip", "install", "pulp"])
    from pulp import *
try:
    import tabulate as tbl
except ImportError:
    print("El módulo tabulate no está instalado. Instalando...")
    import subprocess
    subprocess.check_call(["pip", "install", "tabulate"])
    import tabulate as tbl
try:
    import pandas as pd
except ImportError:
    print("El módulo pandas no está instalado. Instalando...")
    import subprocess
    subprocess.check_call(["pip", "install", "pandas"])
    import pandas as tbl

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
except ImportError:
    print("El módulo openpyxl no está instalado. Instalando...")
    import subprocess
    subprocess.check_call(["pip", "install", "openpyxl"])
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill



#Se definen las funciones
def simplificador_datos(pruva: str) -> list:
    lista_datos_ramo = []
    lista_datos_ramo.append(pruva.find('td', style="font-size:13px;text-align:center;").text.strip())
    lista_datos_ramo.append(pruva.find('td', class_="tooltipInfoCurso").text.strip())
    lista_datos_ramo.append(pruva.find_all('td', style="font-size:13px;text-align:center;")[1].text.strip())
    lista_datos_ramo.append(pruva.find_all('td',style="width:10px;text-align:left;")[1].text.strip())
    lista_datos_ramo.append(pruva.find_all('td',style="font-size:13px;text-align:left;")[1].text.strip())
    lista_datos_ramo.append(pruva.find_all('td',style="font-size:13px;text-align:left;")[3].text.strip())
    lista_datos_ramo.append([pruva.find_all('td',style="font-size:13px;text-align:center;")[4].text.strip(),(pruva.find_all('td', style="font-size:13px;text-align:center;")[2].text.strip())])
    arreglo_2 = []
    for i in range(len(pruva.find("table").text.strip().split("\n"))):
        if pruva.find("table").text.strip().split("\n")[i] == "":
            pass
        else:
            arreglo_2.append(pruva.find("table").text.strip().split("\n")[i])
    lista_datos_ramo.append(arreglo_2)
    return lista_datos_ramo
def key_tercero(lista:list):
    return int(lista[2])
def key_quinta(lista:list):
    return str(lista[4])
def key_sexta(lista:list):
    return str(lista[5])
def opciones_disponibles(sigla_ramo:str, año_actual:str, semestre_actual:str) -> list:
    url_parte_1 = 'https://buscacursos.uc.cl/?cxml_semestre='+año_actual+'-'+semestre_actual+'&cxml_sigla='
    url_parte_2 = '&cxml_nrc=&cxml_nombre=&cxml_categoria=TODOS&cxml_area_fg=T'
    url_parte_3 = 'ODOS&cxml_formato_cur=TODOS&cxml_profesor=&cxml_campus=TODO'
    url_parte_4 = 'S&cxml_unidad_academica=TODOS&cxml_horario_tipo_busqueda=si'
    url_parte_5 ='_tenga&cxml_horario_tipo_busqueda_actividad=TODOS#resultados'
    url =url_parte_1+sigla_ramo+url_parte_2+url_parte_3+url_parte_4+url_parte_5
    response = requests.get(url)
    response_soup = BeautifulSoup(response.text, "html.parser")
    ramos_impar = list(response_soup.find_all('tr', class_="resultadosRowImpar"))
    ramos_par = list(response_soup.find_all('tr', class_="resultadosRowPar"))
    informacion_ramo = ramos_impar + ramos_par
    ramos_ordenados = []
    for elemento in informacion_ramo:
        try:
            ramos_ordenados.append(simplificador_datos(elemento))
        except:
            print(elemento)
    ramos_ordenados.sort(key=key_tercero)
    ramos_ordenados.sort(key=key_sexta)
    ramos_ordenados.sort(key=key_quinta)
    return ramos_ordenados
def opciones_totales(lista_ramos:list, año:str, semestre:str) ->list:
    lista_informacion_total = []
    for ramo in lista_ramos:
        lista_temporal = opciones_disponibles(ramo, año, semestre)
        lista_informacion_total.append(lista_temporal)
    return lista_informacion_total
def calendario_numerico(hora:str) -> list:
    datos = hora.split(":")
    datos[0] = datos[0].split("-")
    datos[1] = datos[1].split(',')
    for i in range(len(datos[1])):
        try: 
            datos[1][i] = int(datos[1][i])
        except:
            pass
    lista_fechas = []
    for i in range(len(datos[0])):
        for j in range(len(datos[1])):
            if datos[0][i] == "L":
                try: 
                    lista_fechas.append((1,datos[1][j]))
                except:
                    pass
            elif datos[0][i] == "M":
                try: 
                    lista_fechas.append((2,datos[1][j]))
                except:
                    pass
            elif datos[0][i] == "W":
                try: 
                    lista_fechas.append((3,datos[1][j]))
                except:
                    pass
            elif datos[0][i] == "J":
                try: 
                    lista_fechas.append((4,datos[1][j]))
                except:
                    pass
            elif datos[0][i] == "V":
                try: 
                    lista_fechas.append((5,datos[1][j]))
                except:
                    pass
            elif datos[0][i] == "S":
                try: 
                    lista_fechas.append((6,datos[1][j]))
                except:
                    pass
    return lista_fechas
def arreglador_fechas(lista_datos_ramos:list) ->None: 
    """Arregla la fecha"""
    for ramo in (range(len(lista_datos_ramos))):
        for seccion in range(len(lista_datos_ramos[ramo])):
            cantidad_elementos = len(lista_datos_ramos[ramo][seccion][7])
            diccionario_modulos = {}
            for i in range(int(cantidad_elementos/3)):
                if lista_datos_ramos[ramo][seccion][7][i*3] == ":":
                    pass
                else:
                    tipo_clase = lista_datos_ramos[ramo][seccion][7][i*3+1]
                    horario_clase = calendario_numerico(lista_datos_ramos[ramo][seccion][7][i*3])
                    if not tipo_clase in diccionario_modulos.keys():
                        diccionario_modulos.update({tipo_clase:horario_clase})
                    else:
                        for j in range(len(horario_clase)):
                            diccionario_modulos[tipo_clase].append(horario_clase[j])
            lista_datos_ramos[ramo][seccion][7] = diccionario_modulos
            if not 'CLAS' in lista_datos_ramos[ramo][seccion][7].keys():
                lista_datos_ramos[ramo][seccion][7].update({'CLAS':[]})
            if not 'TAL' in lista_datos_ramos[ramo][seccion][7].keys():
                lista_datos_ramos[ramo][seccion][7].update({'TAL':[]})
            if not 'LAB' in lista_datos_ramos[ramo][seccion][7].keys():
                lista_datos_ramos[ramo][seccion][7].update({'LAB':[]})
            if not 'AYU' in lista_datos_ramos[ramo][seccion][7].keys():
                lista_datos_ramos[ramo][seccion][7].update({'AYU':[]})
            if not 'TER' in lista_datos_ramos[ramo][seccion][7].keys():
                lista_datos_ramos[ramo][seccion][7].update({'TER':[]})
def extractor_ramos_ofg(tipo_ofg:str) -> list:
    archivo = open(tipo_ofg, "r")
    data = archivo.readlines()
    archivo.close()
    ramos = data[0].strip("\n").split(",")
    return ramos
def numero_dia(numero:int) -> str:
    if numero == 1:
        return 'Lunes'
    elif numero ==2:
        return 'Martes'
    elif numero == 3:
        return 'Miercoles'
    elif numero == 4:
        return 'Jueves'
    elif numero == 5:
        return 'Viernes'
    elif numero == 6:
        return 'Sabado'
def colores_horario(horario:pd.DataFrame):
    if 'CAT' in horario:
        color = 'background-color: rgba(255, 204, 153, 0.5)'
    elif 'TAL' in horario:
        color = 'background-color: rgba(229, 204, 255, 0.5)'
    elif 'LAB' in horario:
        color = 'background-color: rgba(204, 255, 255, 0.5)'
    else:
        color = 'background-color: rgba(255, 255, 255, 0.5)'
    return color
# Realiza una solicitud GET a una URL y se ordenan los datos
año_solicitud = input("En que año estamos: ")
semestre_solicitud = input("En cual semestre estamos:\n    [1]Primer Semestre\n    [2]Segundo Semestre\n    [3]TAV\n    Eleccion: ")
eleccion = []
eleccion_originales = []
creditos_ofg_predeterminado = 10
modulos_antes_almuerzo = 4
sigla_ramo = input("Que ramo quieres agregar (debes ingresar la sigla del ramo, en caso de terminar debes poner 0): ")
while sigla_ramo != "0":
    eleccion.append(sigla_ramo)
    eleccion_originales.append(sigla_ramo)
    sigla_ramo = input("Que ramo quieres agregar (debes ingresar la sigla del ramo, en caso de terminar debes poner 0): ")
#Esta es una eleccion de ramos rapida
#eleccion = ['BIO141C','eyp1113','fis1533','fis0153','iic2233']
#eleccion_originales = ['BIO141C','eyp1113','fis1533','fis0153','iic2233']
ofg_extra = input("Deseas agregar un ofg a tu horario (Este proceso puede demorar):\n    [1]Si\n    [2]No\n    Elección: ")
if ofg_extra == "1":
    tipo_ofg = input('Que tipo de ofg deseas agregar:\n    [1]Ciencias Sociales\n'+
                     '    [2]Artes\n    [3]Teología\n'+
                     '    [4]Ecologia integral y Sustentabilidad\n'+
                     '    [5]Humanidades\n    [6]Salud y Bienestar (10 creditos)\n'+
                     '    [7]Salud y Bienestar (5 creditos)    \nElección: ')
    if tipo_ofg == "1":
        lista_ramos_ofg = extractor_ramos_ofg('ramos de ciencias sociales.txt')
    elif tipo_ofg == "2":
        lista_ramos_ofg = extractor_ramos_ofg('ramos de artes.txt')
    elif tipo_ofg == "3":
        lista_ramos_ofg = extractor_ramos_ofg('ramos de teologia.txt')
    elif tipo_ofg == "4":
        lista_ramos_ofg = extractor_ramos_ofg('ramos de ecologia integral y sustentabilidad.txt')
    elif tipo_ofg == "6":
        lista_ramos_ofg = extractor_ramos_ofg('ramos de salud y bienestar.txt')
    elif tipo_ofg == "5":
        lista_ramos_ofg = extractor_ramos_ofg('ramos de humanidades.txt')
    elif tipo_ofg == '7':
        lista_ramos_ofg = extractor_ramos_ofg('ramos de salud y bienestar.txt')
        creditos_ofg_predeterminado = 5
if ofg_extra != '1':
    resultado = opciones_totales(eleccion, año_solicitud, semestre_solicitud)
    arreglador_fechas(resultado)

    #Se crean los parametros del problema
    #C es el conjunto de numeros que representa a un curso
    cursos_ramos_seleccionados = []
    for ramo in resultado:
        for curso in ramo:
            cursos_ramos_seleccionados.append(int(curso[0]))
    cursos_ramos_seleccionados.sort()
    #R es un conjunto de la forma {1,...,n} donde n es el numero de cursos
    ramos_seleccionados = [i+1 for i in range(len(eleccion))]
    #D es el conjunto de dias {1,...,6}
    dias_disponibles = [i+1 for i in range(6)]
    #T es el conjunto de modulos {1,...,9} dentro de un dia
    modulos_disponibles = [i+1 for i in range(9)]

    #Se calculan los Datos
    #curso_ramo[(i,j)] es un dato binario que es 1 ssi el curso i pertenece al ramo j
    lista_conjunto_cursos = []
    for ramo in resultado:
        conjunto_cursos = set()
        for curso in ramo:
            conjunto_cursos.add(int(curso[0]))
        lista_conjunto_cursos.append(conjunto_cursos)
    curso_ramo = {}
    for i in range(len(cursos_ramos_seleccionados)):
        for j in ramos_seleccionados:
            if cursos_ramos_seleccionados[i] in lista_conjunto_cursos[j-1]:
                curso_ramo.update({(cursos_ramos_seleccionados[i],j):1})
            else:
                curso_ramo.update({(cursos_ramos_seleccionados[i],j):0})
    #h[(i,j,k)] es un dato que es 1 ssi el curso i es impartido el dia j el modulo k
    lista_curso_fechas = []
    for ramo in resultado:
        for curso in ramo:
            lista_curso_fechas.append([int(curso[0]), curso[7]['CLAS'] + curso[7]['TAL'] + curso[7]['TER']])
    catedra_dia_modulo = {}
    for i in range(len(cursos_ramos_seleccionados)):
        for j in dias_disponibles:
            for k in modulos_disponibles:
                for curso in lista_curso_fechas:
                    if curso[0] == cursos_ramos_seleccionados[i] and (j,k) in curso[1]:
                        catedra_dia_modulo.update({(cursos_ramos_seleccionados[i],j,k):1})
                        break
                else:
                    catedra_dia_modulo.update({(cursos_ramos_seleccionados[i],j,k):0})

    lista_ayudantia_fechas = []
    for ramo in resultado:
        for curso in ramo:
            lista_ayudantia_fechas.append([int(curso[0]), curso[7]['AYU']])
    ayudantia_dia_modulo = {}
    for i in range(len(cursos_ramos_seleccionados)):
        for j in dias_disponibles:
            for k in modulos_disponibles:
                for curso in lista_ayudantia_fechas:
                    if curso[0] == cursos_ramos_seleccionados[i] and (j,k) in curso[1]:
                        ayudantia_dia_modulo.update({(cursos_ramos_seleccionados[i],j,k):1})
                        break
                else:
                    ayudantia_dia_modulo.update({(cursos_ramos_seleccionados[i],j,k):0})
    lista_laboratorio_curso = []
    for ramo in resultado:
        for curso in ramo:
            lista_laboratorio_curso.append([int(curso[0]), curso[7]['LAB']])
    laboratorio_dia_modulo = {}
    for i in range(len(cursos_ramos_seleccionados)):
        for j in dias_disponibles:
            for k in modulos_disponibles:
                for curso in lista_laboratorio_curso:
                    if curso[0] == cursos_ramos_seleccionados[i] and (j,k) in curso[1]:
                        laboratorio_dia_modulo.update({(cursos_ramos_seleccionados[i],j,k):1})
                        break
                else:
                    laboratorio_dia_modulo.update({(cursos_ramos_seleccionados[i],j,k):0})

    #Se crea un diccionario que contiene el numero de un curso, y devuelve el nombre con la seccion
    informacion_curso = {}
    for ramo in resultado:
        for curso in ramo:
            informacion_curso.update({int(curso[0]):[curso[0], curso[4], curso[2],curso[5],curso[7]]})

    #Se configura el optimizador (Pulp en este caso)

    # Crear el problema
    horario = LpProblem("Horario_universitario", LpMinimize)


    x = LpVariable.dicts('x', cursos_ramos_seleccionados, cat='Binary') #Variable que es 1 ssi se toma el el ramo i
    y = LpVariable.dicts('y', dias_disponibles, cat='Binary') #Variable que es 0 ssi el dia i es libre
    z = LpVariable.dicts('z', dias_disponibles) #Variable que indica la cantidad de ramos en un dia
    z_1 = LpVariable.dicts('z_1', dias_disponibles) #Variable auxiliar que menciona el primer modulo del dia
    z_2 = LpVariable.dicts('z_2', dias_disponibles) #Variable auxiliar que indica el ultimo modulo del dia

    # Función objetivo
    horario += lpSum([y[j]*len(ramos_seleccionados) + z_2[j] - z[j] 
                    for j in dias_disponibles ]) + lpSum([catedra_dia_modulo[(i, j, k)]*x[i]*k + 
                    laboratorio_dia_modulo[(i, j, k)]*x[i]*(k/2)  + 
                    ayudantia_dia_modulo[(i, j, k)]*x[i]*(k/10) + 
                    y[j]*len(ramos_seleccionados) + z_2[j] - z[j] 
                    if k > modulos_antes_almuerzo else 0
                    for i in cursos_ramos_seleccionados 
                    for j in dias_disponibles 
                    for k in modulos_disponibles])

    # Restricción 

    for j in ramos_seleccionados:
        horario += lpSum(curso_ramo[(i, j)]*x[i] for i in cursos_ramos_seleccionados) == 1

    for j in dias_disponibles:
        for k in modulos_disponibles:
            horario += lpSum(catedra_dia_modulo[(i, j, k)]*x[i] + 
                            laboratorio_dia_modulo[(i, j, k)]*x[i]
                            for i in cursos_ramos_seleccionados) <= 1
            horario += lpSum(catedra_dia_modulo[(i, j, k)]*x[i] + 
                            ayudantia_dia_modulo[(i, j, k)]*x[i] + 
                            laboratorio_dia_modulo[(i, j, k)]*x[i] 
                            for i in cursos_ramos_seleccionados) <= 2
            horario += lpSum(catedra_dia_modulo[(i, j, k)]*x[i]*k 
                            for i in cursos_ramos_seleccionados) <= z_2[j]
            horario += lpSum(catedra_dia_modulo[(i, j, k)]*x[i]*k 
                            for i in cursos_ramos_seleccionados) >= z_1[j]
        horario += lpSum(catedra_dia_modulo[(i, j, k)]*x[i] 
                        for i in cursos_ramos_seleccionados 
                        for k in modulos_disponibles) <= y[j] * len(modulos_disponibles)
        horario += lpSum(catedra_dia_modulo[(i, j, k)]*x[i] 
                        for i in cursos_ramos_seleccionados 
                        for k in modulos_disponibles) >= y[j]
        horario += lpSum(catedra_dia_modulo[(i, j, k)]*x[i] 
                        for i in cursos_ramos_seleccionados 
                        for k in modulos_disponibles) == z[j]


    # Resolver el problema
    horario.solve()

else:
    eleccion += lista_ramos_ofg
    resultado = opciones_totales(eleccion, año_solicitud, semestre_solicitud)
    for ramo in list(reversed(range(len(resultado))))[:-len(eleccion_originales)]:
        if not resultado[ramo] or 'Selección' in resultado[ramo][0][4] or ' II' in resultado[ramo][0][4] or len(resultado[ramo][0]) != 8 or resultado[ramo][0][5] != 'San Joaquín' or 'Seleción' in resultado[ramo][0][4] or not resultado[ramo][0][6][1] == str(creditos_ofg_predeterminado): # or int(resultado[ramo][0][6][0]) == 0
            resultado.pop(ramo)
    arreglador_fechas(resultado)
    #Se crean los parametros del problema
    #C es el conjunto de numeros que representa a un curso
    cursos_ramos_seleccionados = []
    for ramo in resultado:
        for curso in ramo:
            cursos_ramos_seleccionados.append(int(curso[0]))
    cursos_ramos_seleccionados.sort()
    #R es un conjunto de la forma {1,...,n} donde n es el numero de cursos
    ramos_seleccionados = [i+1 for i in range(len(resultado))]
    ramos_seleccionados_originales = [i+1 for i in range(len(eleccion_originales))]
    #D es el conjunto de dias {1,...,6}
    dias_disponibles = [i+1 for i in range(6)]
    #T es el conjunto de modulos {1,...,9} dentro de un dia
    modulos_disponibles = [i+1 for i in range(9)]

    #Se calculan los Datos
    #curso_ramo[(i,j)] es un dato binario que es 1 ssi el curso i pertenece al ramo j
    lista_conjunto_cursos = []
    for ramo in resultado:
        conjunto_cursos = set()
        for curso in ramo:
            conjunto_cursos.add(int(curso[0]))
        lista_conjunto_cursos.append(conjunto_cursos)
    curso_ramo = {}
    for i in range(len(cursos_ramos_seleccionados)):
        for j in ramos_seleccionados:
            if cursos_ramos_seleccionados[i] in lista_conjunto_cursos[j-1]:
                curso_ramo.update({(cursos_ramos_seleccionados[i],j):1})
            else:
                curso_ramo.update({(cursos_ramos_seleccionados[i],j):0})
    #h[(i,j,k)] es un dato que es 1 ssi el curso i es impartido el dia j el modulo k
    lista_curso_fechas = []
    for ramo in resultado:
        for curso in ramo:
            lista_curso_fechas.append([int(curso[0]), curso[7]['CLAS'] + curso[7]['TAL'] + curso[7]['TER']])
    catedra_dia_modulo = {}
    for i in range(len(cursos_ramos_seleccionados)):
        for j in dias_disponibles:
            for k in modulos_disponibles:
                for curso in lista_curso_fechas:
                    if curso[0] == cursos_ramos_seleccionados[i] and (j,k) in curso[1]:
                        catedra_dia_modulo.update({(cursos_ramos_seleccionados[i],j,k):1})
                        break
                else:
                    catedra_dia_modulo.update({(cursos_ramos_seleccionados[i],j,k):0})

    lista_ayudantia_fechas = []
    for ramo in resultado:
        for curso in ramo:
            lista_ayudantia_fechas.append([int(curso[0]), curso[7]['AYU']])
    ayudantia_dia_modulo = {}
    for i in range(len(cursos_ramos_seleccionados)):
        for j in dias_disponibles:
            for k in modulos_disponibles:
                for curso in lista_ayudantia_fechas:
                    if curso[0] == cursos_ramos_seleccionados[i] and (j,k) in curso[1]:
                        ayudantia_dia_modulo.update({(cursos_ramos_seleccionados[i],j,k):1})
                        break
                else:
                    ayudantia_dia_modulo.update({(cursos_ramos_seleccionados[i],j,k):0})
    lista_laboratorio_curso = []
    for ramo in resultado:
        for curso in ramo:
            lista_laboratorio_curso.append([int(curso[0]), curso[7]['LAB']])
    laboratorio_dia_modulo = {}
    for i in range(len(cursos_ramos_seleccionados)):
        for j in dias_disponibles:
            for k in modulos_disponibles:
                for curso in lista_laboratorio_curso:
                    if curso[0] == cursos_ramos_seleccionados[i] and (j,k) in curso[1]:
                        laboratorio_dia_modulo.update({(cursos_ramos_seleccionados[i],j,k):1})
                        break
                else:
                    laboratorio_dia_modulo.update({(cursos_ramos_seleccionados[i],j,k):0})

    #Se crea un diccionario que contiene el numero de un curso, y devuelve el nombre con la seccion 
    informacion_curso = {}
    for ramo in resultado:
        for curso in ramo:
            informacion_curso.update({int(curso[0]):[curso[0], curso[4], curso[2],curso[5],curso[7]]})


    #Se configura el optimizador (Pulp en este caso)
    #Modifica para que elija un solo ofg y los primeros ramos

    # Crear el problema
    horario = LpProblem("Horario_universitario", LpMinimize)


    x = LpVariable.dicts('x', cursos_ramos_seleccionados, cat='Binary') #Variable que es 1 ssi se toma el el ramo i
    y = LpVariable.dicts('y', dias_disponibles, cat='Binary') #Variable que es 0 ssi el dia i es libre
    z = LpVariable.dicts('z', dias_disponibles) #Variable que indica la cantidad de ramos en un dia
    z_1 = LpVariable.dicts('z_1', dias_disponibles) #Variable auxiliar que menciona el primer modulo del dia
    z_2 = LpVariable.dicts('z_2', dias_disponibles) #Variable auxiliar que indica el ultimo modulo del dia

    # Función objetivo
    horario += lpSum([y[j]*len(ramos_seleccionados) + z_2[j] -z_1[j] - z[j] 
                    for j in dias_disponibles ]) + lpSum([catedra_dia_modulo[(i, j, k)]*x[i]*k + 
                    laboratorio_dia_modulo[(i, j, k)]*x[i]*(k/2)  + 
                    ayudantia_dia_modulo[(i, j, k)]*x[i]*(k/10) + 
                    y[j]*len(ramos_seleccionados) + z_2[j] - z[j] 
                    if k > modulos_antes_almuerzo else 0
                    for i in cursos_ramos_seleccionados 
                    for j in dias_disponibles 
                    for k in modulos_disponibles])

    # Restricción 

    for j in ramos_seleccionados_originales:
        horario += lpSum(curso_ramo[(i, j)]*x[i] for i in cursos_ramos_seleccionados) == 1
    
    horario += lpSum(x[i] for i in cursos_ramos_seleccionados) == (len(ramos_seleccionados_originales)+1)

    for j in dias_disponibles:
        for k in modulos_disponibles:
            horario += lpSum(catedra_dia_modulo[(i, j, k)]*x[i] + 
                            laboratorio_dia_modulo[(i, j, k)]*x[i]
                            for i in cursos_ramos_seleccionados) <= 1
            horario += lpSum(catedra_dia_modulo[(i, j, k)]*x[i] + 
                            ayudantia_dia_modulo[(i, j, k)]*x[i] + 
                            laboratorio_dia_modulo[(i, j, k)]*x[i] 
                            for i in cursos_ramos_seleccionados) <= 2
            horario += lpSum(catedra_dia_modulo[(i, j, k)]*x[i]*k 
                            for i in cursos_ramos_seleccionados) <= z_2[j]
            horario += lpSum(catedra_dia_modulo[(i, j, k)]*x[i]*k 
                            for i in cursos_ramos_seleccionados) >= z_1[j]
        horario += lpSum(catedra_dia_modulo[(i, j, k)]*x[i] +
                        laboratorio_dia_modulo[(i, j, k)]*x[i] 
                        for i in cursos_ramos_seleccionados 
                        for k in modulos_disponibles) <= y[j] * len(modulos_disponibles)
        horario += lpSum(catedra_dia_modulo[(i, j, k)]*x[i] +
                        laboratorio_dia_modulo[(i, j, k)]*x[i] 
                        for i in cursos_ramos_seleccionados 
                        for k in modulos_disponibles) >= y[j]
        horario += lpSum(catedra_dia_modulo[(i, j, k)]*x[i] 
                        for i in cursos_ramos_seleccionados 
                        for k in modulos_disponibles) == z[j]


    # Resolver el problema
    horario.solve()


#Se crea el display de los datos
encabezados = ["NRC","Nombre","Seccion","Campus","Horarios"]
filas = []
for i in cursos_ramos_seleccionados:
    if x[i].varValue> 0:
        filas.append(informacion_curso[i])
# Crear la tabla
tabla = tbl.tabulate(filas, headers=encabezados, tablefmt='fancy_grid')

# Mostrar la tabla
print(tabla)
print(f"Estado de la solucion {LpStatus[horario.status]}")


diccionario_pruva = {'Horas':['8:20-9:30','9:40-10:50','11:00-12:10','12:20-13:30','14:50-16:00','16:10-17:20','17:30-18:40'],
                     'Lunes':['','','','','','',''],
                     'Martes':['','','','','','',''],
                     'Miercoles':['','','','','','',''],
                     'Jueves':['','','','','','',''],
                     'Viernes':['','','','','','',''],
                     'Sabado':['','','','','','','']}
horario_final = pd.DataFrame(diccionario_pruva)
for clase in filas:
    for tipo, hora in clase[4].items():
        if not hora:
            continue
        for modulo in hora:
            horario_final.at[modulo[1]-1,numero_dia(modulo[0])] = f'{clase[1]} ({tipo})'


# Crear un nuevo libro de Excel
wb = Workbook()
ws = wb.active

# Convertir el DataFrame en una lista de listas para escribirlo en el archivo de Excel
datos = [horario_final.columns.tolist()] + horario_final.values.tolist()

# Escribir los datos en el archivo de Excel y aplicar estilos de relleno
for row_idx, row in enumerate(datos, start=1):
    for col_idx, value in enumerate(row, start=1):
        ws.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 1:  # Estilos para la fila de encabezados
            ws.cell(row=row_idx, column=col_idx).fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')  
        elif col_idx == 1:  # Estilos para la fila de encabezados
            ws.cell(row=row_idx, column=col_idx).fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
        elif 'CLAS' in value:
            ws.cell(row=row_idx, column=col_idx).fill = PatternFill(start_color='FFCC99', end_color='FFCC99', fill_type='solid')  
        elif 'TAL' in value:
            ws.cell(row=row_idx, column=col_idx).fill = PatternFill(start_color='E5CCFF', end_color='E5CCFF', fill_type='solid')  
        elif 'LAB' in value:
            ws.cell(row=row_idx, column=col_idx).fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')  
        elif 'AYU' in value:
            ws.cell(row=row_idx, column=col_idx).fill = PatternFill(start_color='99FF99', end_color='99FF99', fill_type='solid')  
        elif 'TER' in value:
            ws.cell(row=row_idx, column=col_idx).fill = PatternFill(start_color='FFCCE5', end_color='FFCCE5', fill_type='solid')  
        else:
            ws.cell(row=row_idx, column=col_idx).fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')  

for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2
    ws.column_dimensions[column_letter].width = adjusted_width

# Guardar el archivo de Excel
wb.save('Horario.xlsx')

print("Se a creado/modificad un excel con el horario llamada Hoario.xsl")
#final = input('Presione Enter para finalizar')
#[eyp1113,fis1533,fis0153,iic1005,iic2233] print